import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import PatternFill, Font

from actions import perform_action
from verify import verify_expected_text
from utils import set_hyperlink


added_column_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
header_font = Font(bold=True, color="000000") 

def init_driver(url: str, browser: str): 
    if browser == "Chrome":
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=options)
    elif browser == "Edge":
        driver = webdriver.Edge()
    elif browser == "Safari":
        driver = webdriver.Safari()
    else:
        raise ValueError(f"未対応のブラウザです : {browser}")

    driver.get(url)
    return driver

def load_excel(scenario_file: str):
    # Excel読み込み
    wb = openpyxl.load_workbook(scenario_file)
    ws = wb.active
    return wb, ws

def prepare_columns(ws):
    header = [cell.value for cell in ws[1]]
    idx = {
        "test_id": header.index("Test Case ID"),
        "action": header.index("Action"),
        "selector": header.index("Selector"),
        "value": header.index("Value"),
        "expected": header.index("Expected Result"),
        "result": len(header),
        "error": len(header) + 1,
        "screenshot": len(header) + 2,
        "verify": len(header) + 3
    }
    cell = ws.cell(row=1, column=idx["result"] + 1)
    cell.value = "Status"
    cell.fill = added_column_fill
    cell.font = header_font
  
    cell = ws.cell(row=1, column=idx["error"] + 1)
    cell.value = "Error"
    cell.fill = added_column_fill
    cell.font = header_font

    cell = ws.cell(row=1, column=idx["screenshot"] + 1)
    cell.value = "Screenshot"
    cell.fill = added_column_fill
    cell.font = header_font    

    cell = ws.cell(row=1, column=idx["verify"] + 1)
    cell.value = "Verify Screenshot"
    cell.fill = added_column_fill
    cell.font = header_font    

    return idx

def run_test_scenario(
    url: str,
    scenario_file: str,
    wait_time: int,
    use_regex: bool,
    verify_enabled: bool,
    screenshot_dir: str,
    browser: str,
    log_func=print 
) -> str:
    try: 
        driver = init_driver(url, browser)
        log_func(f"[INFO] ブラウザ初期化: {browser}")

        wb, ws = load_excel(scenario_file)
        log_func(f"[INFO] Excel読み込み: {scenario_file}")

        idx = prepare_columns(ws)      

        # Excel読み込み・ループ処理
        log_func(f"[INFO] テスト開始: {url}")
        for row in ws.iter_rows(min_row=2, values_only=False):
            test_id = row[idx["test_id"]].value
            if not test_id or str(test_id).strip() == "":
                continue
          
            verify_screenshot = None  # 未定義に備えた初期化

            action = row[idx["action"]].value
            selector = row[idx["selector"]].value
            value = row[idx["value"]].value
            expected = row[idx["expected"]].value

            log_func(f"[TEST] Case: {test_id} | Action: {action} | Selector: {selector}")
            success, error, screenshot = perform_action(driver, action, selector, value,
                                                        wait_time, screenshot_dir, log_func)

            if success and expected and verify_enabled:
                verified, verify_error, verify_screenshot = verify_expected_text(driver, expected, wait_time,
                                                                                 use_regex, screenshot_dir)
                if not verified:
                    success = False
                    error = verify_error
                    row[idx["verify"]].value = verify_screenshot
                else:
                    row[idx["verify"]].value = ""

            row[idx["result"]].value = "Success" if success else "Failed"
            row[idx["error"]].value = error
            set_hyperlink(row[idx["screenshot"]], screenshot)
            set_hyperlink(row[idx["verify"]], verify_screenshot)

            if success:
                log_func(f"[PASS] {test_id} 成功")
            else:
                log_func(f"[FAIL] {test_id} 失敗 | Error: {error}")

        # レポート保存
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        report_path = f"./reports/test_report_{timestamp}.xlsx"
        wb.save(report_path)
        driver.quit()
        print(f"[INFO] テスト完了。レポート保存済み: {report_path}")
        return report_path

    except Exception as e:
        print(f"エラー発生: {e}")
        return "エラーが発生しました"